"""Read test data (username, password, products) from Excel."""

import os

from openpyxl import load_workbook

from utilities.readProperties import ReadConfig


class ReadExcel:
    """Excel file: Data/test_data.xlsx
    Sheet 'Users'    → username, password
    Sheet 'Products' → name, price, add_to_cart_id
    """

    _project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    USERS_SHEET = "Users"
    PRODUCTS_SHEET = "Products"

    @classmethod
    def _excel_path(cls):
        """Build full path to Excel file from config.ini."""
        path = ReadConfig.getExcelPath()
        if not os.path.isabs(path):
            path = os.path.join(cls._project_root, path)
        return path

    @classmethod
    def _get_sheet(cls, sheet_name):
        """Open workbook and return the requested sheet."""
        excel_path = cls._excel_path()
        try:
            workbook = load_workbook(excel_path, read_only=True, data_only=True)
        except PermissionError as exc:
            raise PermissionError(
                f"Cannot read '{excel_path}'. Close Excel and run again."
            ) from exc
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Sheet '{sheet_name}' not found.")
        return workbook, workbook[sheet_name]

    @classmethod
    def get_user(cls, username=None):
        """Return one user dict from Users sheet. Uses default username if none given."""
        workbook, sheet = cls._get_sheet(cls.USERS_SHEET)
        if username is None:
            username = ReadConfig.getDefaultUsername()
        for row in sheet.iter_rows(min_row=2, values_only=True):  # row 1 = header
            if row[0] and row[1] and str(row[0]) == username:
                workbook.close()
                return {"username": str(row[0]), "password": str(row[1])}
        workbook.close()
        raise ValueError(f"Username '{username}' not found.")

    @classmethod
    def get_username(cls, username=None):
        return cls.get_user(username)["username"]

    @classmethod
    def get_password(cls, username=None):
        return cls.get_user(username)["password"]

    @classmethod
    def get_product(cls, row_index=0):
        """Return one product by row index (0 = first product in Excel)."""
        workbook, sheet = cls._get_sheet(cls.PRODUCTS_SHEET)
        products = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, price, add_to_cart_id = row[0], row[1], row[2]
            if name and add_to_cart_id:
                products.append({
                    "name": str(name),
                    "price": float(price),
                    "add_to_cart_id": str(add_to_cart_id),
                })
        workbook.close()
        if not products:
            raise ValueError("No products found in Excel.")
        if row_index < 0 or row_index >= len(products):
            raise IndexError(f"Product row {row_index} out of range.")
        return products[row_index]
